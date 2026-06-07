from csv import DictReader
from pathlib import Path
import time

from postgresql.postgres_setup import (
    DEFAULT_CONTENT_TABLE,
    DEFAULT_MOVIE_METADATA_TABLE,
    DEFAULT_TV_METADATA_TABLE,
    DEFAULT_USER_ACCOUNT_TABLE,
    DEFAULT_USER_CONTENT_TABLE,
    count_rows,
    get_connection,
    wait_for_postgres,
)


ROOT_DIR = Path(__file__).resolve().parent.parent
TV_CSV_PATH = ROOT_DIR / "netflix_tv_shows_detailed_up_to_2025.csv"
MOVIE_CSV_PATH = ROOT_DIR / "netflix_movies_detailed_up_to_2025.csv"
USER_XLSX_PATH = ROOT_DIR / "user_data.xlsx"


def wait_for_startup_import(timeout_seconds: int = 60) -> None:
    deadline = time.time() + timeout_seconds

    while time.time() < deadline:
        wait_for_postgres()
        with get_connection() as connection:
            content_total = count_rows(connection, DEFAULT_CONTENT_TABLE)
            tv_total = count_rows(connection, DEFAULT_TV_METADATA_TABLE)
            movie_total = count_rows(connection, DEFAULT_MOVIE_METADATA_TABLE)
            user_content_total = count_rows(connection, DEFAULT_USER_CONTENT_TABLE)
            user_account_total = count_rows(connection, DEFAULT_USER_ACCOUNT_TABLE)

        if all(total > 0 for total in [content_total, tv_total, movie_total, user_content_total, user_account_total]):
            return

        time.sleep(1)

    raise AssertionError("Startup import did not finish loading dataset rows before timeout.")


def test_startup_import_loaded_dataset():
    with TV_CSV_PATH.open(newline="", encoding="utf-8") as tv_csv:
        tv_rows = list(DictReader(tv_csv))
    with MOVIE_CSV_PATH.open(newline="", encoding="utf-8") as movie_csv:
        movie_rows = list(DictReader(movie_csv))
    from openpyxl import load_workbook

    workbook = load_workbook(USER_XLSX_PATH, read_only=True, data_only=True)
    try:
        user_content_rows = list(workbook["user_content"].iter_rows(values_only=True))[1:]
        user_account_rows = list(workbook["user_account"].iter_rows(values_only=True))[1:]
    finally:
        workbook.close()

    expected_content_total = len(
        {row["show_id"] for row in tv_rows + movie_rows if row.get("show_id")}
    )
    expected_tv_total = len({row["show_id"] for row in tv_rows if row.get("show_id")})
    expected_movie_total = len(
        {row["show_id"] for row in movie_rows if row.get("show_id")}
    )
    expected_user_content_total = len(
        {
            (int(row[0]), int(row[1]))
            for row in user_content_rows
            if row[0] is not None and row[1] is not None
        }
    )
    expected_user_account_total = len(
        {
            int(row[0])
            for row in user_account_rows
            if row[0] is not None and row[1] is not None
        }
    )

    wait_for_startup_import()
    with get_connection() as connection:
        content_total = count_rows(connection, DEFAULT_CONTENT_TABLE)
        tv_total = count_rows(connection, DEFAULT_TV_METADATA_TABLE)
        movie_total = count_rows(connection, DEFAULT_MOVIE_METADATA_TABLE)
        user_content_total = count_rows(connection, DEFAULT_USER_CONTENT_TABLE)
        user_account_total = count_rows(connection, DEFAULT_USER_ACCOUNT_TABLE)
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT password FROM {DEFAULT_USER_ACCOUNT_TABLE} WHERE user_id = 1"
            )
            user_account_row = cursor.fetchone()

    assert content_total == expected_content_total
    assert tv_total == expected_tv_total
    assert movie_total == expected_movie_total
    assert user_content_total == expected_user_content_total
    assert user_account_total == expected_user_account_total
    assert (
        user_account_row["password"]
        == "f02302bcd9e4cd85e9d49756ec32030460f673205ac2f258f8433fc11d9fc3db"
    )
