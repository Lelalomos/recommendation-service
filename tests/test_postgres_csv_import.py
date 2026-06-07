from pathlib import Path
from csv import DictReader
from openpyxl import load_workbook
import pytest

from postgresql.postgres_setup import (
    DEFAULT_CONTENT_TABLE,
    DEFAULT_MOVIE_METADATA_TABLE,
    DEFAULT_TV_METADATA_TABLE,
    DEFAULT_USER_ACCOUNT_TABLE,
    DEFAULT_USER_CONTENT_TABLE,
    count_rows,
    create_import_tables,
    find_content_by_title,
    get_connection,
    import_csv_dataset,
    reset_import_tables,
    wait_for_postgres,
)


ROOT_DIR = Path(__file__).resolve().parent.parent
TV_CSV_PATH = ROOT_DIR / "netflix_tv_shows_detailed_up_to_2025.csv"
MOVIE_CSV_PATH = ROOT_DIR / "netflix_movies_detailed_up_to_2025.csv"
USER_XLSX_PATH = ROOT_DIR / "user_data.xlsx"


def setup_function():
    wait_for_postgres()
    with get_connection() as connection:
        reset_import_tables(connection)
        create_import_tables(connection)


def test_import_csv_dataset_loads_expected_row_counts():
    with TV_CSV_PATH.open(newline="", encoding="utf-8") as tv_csv:
        tv_rows = list(DictReader(tv_csv))
    with MOVIE_CSV_PATH.open(newline="", encoding="utf-8") as movie_csv:
        movie_rows = list(DictReader(movie_csv))
    workbook = load_workbook(USER_XLSX_PATH, read_only=True, data_only=True)
    try:
        user_content_rows = list(workbook["user_content"].iter_rows(values_only=True))[1:]
        user_account_rows = list(workbook["user_account"].iter_rows(values_only=True))[1:]
    finally:
        workbook.close()
    unique_show_ids = {row["show_id"] for row in tv_rows + movie_rows if row.get("show_id")}
    unique_tv_show_ids = {row["show_id"] for row in tv_rows if row.get("show_id")}
    unique_movie_show_ids = {row["show_id"] for row in movie_rows if row.get("show_id")}
    unique_user_content = {
        (int(row[0]), int(row[1]))
        for row in user_content_rows
        if row[0] is not None and row[1] is not None
    }
    unique_user_accounts = {
        int(row[0]): (str(row[1]).strip(), str(row[2]).strip())
        for row in user_account_rows
        if row[0] is not None and row[1] is not None and row[2] is not None
    }

    with get_connection() as connection:
        results = import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)

        content_total = count_rows(connection, DEFAULT_CONTENT_TABLE)
        tv_total = count_rows(connection, DEFAULT_TV_METADATA_TABLE)
        movie_total = count_rows(connection, DEFAULT_MOVIE_METADATA_TABLE)
        user_content_total = count_rows(connection, DEFAULT_USER_CONTENT_TABLE)
        user_account_total = count_rows(connection, DEFAULT_USER_ACCOUNT_TABLE)

    assert results[DEFAULT_CONTENT_TABLE] == content_total
    assert results[DEFAULT_TV_METADATA_TABLE] == tv_total
    assert results[DEFAULT_MOVIE_METADATA_TABLE] == movie_total
    assert results[DEFAULT_USER_CONTENT_TABLE] == user_content_total
    assert results[DEFAULT_USER_ACCOUNT_TABLE] == user_account_total
    assert content_total == len(unique_show_ids)
    assert tv_total == len(unique_tv_show_ids)
    assert movie_total == len(unique_movie_show_ids)
    assert user_content_total == len(unique_user_content)
    assert user_account_total == len(unique_user_accounts)


def test_tv_metadata_uses_null_budget_and_revenue():
    with get_connection() as connection:
        import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT budget, revenue
                FROM {DEFAULT_TV_METADATA_TABLE}
                WHERE show_id IS NOT NULL
                LIMIT 1
                """
            )
            row = cursor.fetchone()

    assert row["budget"] is None
    assert row["revenue"] is None


def test_movie_metadata_keeps_budget_and_revenue_values():
    with get_connection() as connection:
        import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT budget, revenue
                FROM {DEFAULT_MOVIE_METADATA_TABLE}
                WHERE budget IS NOT NULL OR revenue IS NOT NULL
                LIMIT 1
                """
            )
            row = cursor.fetchone()

    assert row is not None


def test_import_csv_dataset_is_repeatable():
    with get_connection() as connection:
        first = import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        second = import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        content_total = count_rows(connection, DEFAULT_CONTENT_TABLE)
        user_account_total = count_rows(connection, DEFAULT_USER_ACCOUNT_TABLE)

    assert first == second
    assert content_total == second[DEFAULT_CONTENT_TABLE]
    assert user_account_total == second[DEFAULT_USER_ACCOUNT_TABLE]


@pytest.mark.parametrize("lookup_title", ["ironman", "IRON MAN", "iron-man", " iron   man "])
def test_find_content_by_title_matches_normalized_title(lookup_title):
    with get_connection() as connection:
        import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        exact_match = find_content_by_title(connection, "Iron Man")
        normalized_match = find_content_by_title(connection, lookup_title)

    assert exact_match is not None
    assert normalized_match is not None
    assert str(exact_match["show_id"]) == "45418"
    assert str(normalized_match["show_id"]) == "45418"
    assert normalized_match["title"] == "Iron Man"


def test_user_workbook_tables_load_expected_values():
    with get_connection() as connection:
        import_csv_dataset(connection, TV_CSV_PATH, MOVIE_CSV_PATH, USER_XLSX_PATH)
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT user_id, show_id FROM {DEFAULT_USER_CONTENT_TABLE} ORDER BY user_id, show_id LIMIT 1"
            )
            user_content_row = cursor.fetchone()
            cursor.execute(
                f"SELECT user_id, username, password FROM {DEFAULT_USER_ACCOUNT_TABLE} ORDER BY user_id LIMIT 1"
            )
            user_account_row = cursor.fetchone()

    assert user_content_row["user_id"] == 1
    assert user_content_row["show_id"] == 27205
    assert user_account_row["user_id"] == 1
    assert user_account_row["username"] == "lelalomos"
    assert (
        user_account_row["password"]
        == "f02302bcd9e4cd85e9d49756ec32030460f673205ac2f258f8433fc11d9fc3db"
    )
